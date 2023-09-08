import openpyxl as excel
import os
import datetime

cd = os.getcwd()
cd = cd + "\\Python project\\Excel project\\Year_age.xlsx"

book = excel.load_workbook(cd)
sheet = book.worksheets[0]

#年度を得る
thisyear = datetime.date.today().year

for i in range(80):
    sheet.cell(row=i+1, column=1, value=str(thisyear-i) +"年")
    sheet.cell(row=i+1, column=2, value=str(i) +"歳")


base_year = datetime.date.today().year -10

for j in range(50):
    year = base_year +j
    s1 = year -7 #4/2以降生まれ
    s2 = year -6 #早生まれ
    sf = "{}年4/2～{}年4/1生まれの人".format(s1,s2)
    sheet.cell(row=j+1, column=3, value=str(year) +"年度")
    sheet.cell(row=j+1, column=4, value=sf)

book.save(cd)