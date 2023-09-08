import openpyxl as excel
import os
from datetime import datetime

cd = os.getcwd()
cd = cd + "\\Python project\\Excel project\\wareki.xlsx"

book = excel.load_workbook(cd)
sheet = book.worksheets[0]

wareki_table = [
    {"name":"明治", "start":1868, "end":1912},
    {"name":"大正", "start":1912, "end":1926},
    {"name":"昭和", "start":1926, "end":1989},
    {"name":"平成", "start":1989, "end":2019},
    {"name":"令和", "start":2019, "end":9999}
]

#西暦から和暦へ
def seireki_wareki(year):
    for w in wareki_table:
        if w["start"] <= year < w["end"]:
            y = str(year - w["start"] +1) +"年"
            if y =="1年": y ="元年"
            return w["name"] + y
    
    return "不明"


#見出しをつける
sheet["A1"] = "西暦"
sheet["B1"] = "和暦"

start_year = 1870
for i in range(80):
    year = start_year +i+i*2
    wa = seireki_wareki(year)
    sheet.cell(row=i+2, column=1, value=str(year) +"年")
    sheet.cell(row=i+2, column=2, value=wa)
    sheet.cell(row=i+2, column=3, value="=text(A" +str(i+2)+ ",\"ggge年m月d日\")")
    print(str(year) +"-"+ wa + sheet.cell(row=i+2,column=3).value)

book.save(cd)