import openpyxl as excel
import os

cd  = os.getcwd()
cd = cd + "\\Python project\\Excel project\\Hello.xlsx"

book = excel.load_workbook(cd)
sheet = book.worksheets[0]

sheet["B3"] = "B3セルです"
#sheet.cell(row=1, column=2, value="B3セルです")
#cell = sheet.cell(row=1, column=2)
#cell.value = "B3セルです"

#シート内20行目から30列目までを一括クリア
for row in sheet.iter_rows(max_row=20,max_col=30):
    for cell in row:
        cell.value = None

#九九表を作成
for i in range(1,10):
    for j in range(1,10):
        sheet.cell(row=i, column=j, value=i*j)

#セル番号を100×100に表示
for i in range(1,101):
    for j in range(1,101):
        cell = sheet.cell(i, j)
        cell.value = cell.coordinate #セル番号を引用

book.save(cd)